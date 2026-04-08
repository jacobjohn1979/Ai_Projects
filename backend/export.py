"""
export.py — Statistics export to Excel/CSV
Monthly reports, case exports, flag analysis.
"""
import os
import io
import csv
import json
import logging
from datetime import datetime, timedelta
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv

load_dotenv()
log = logging.getLogger("fraud_detect.export")

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://fraud:fraudpass@postgres:5432/fraud_detect")
engine       = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    log.warning("openpyxl not installed — Excel export unavailable")


def export_cases_csv(days: int = 30, risk_level: str = "") -> bytes:
    """Export screening cases as CSV."""
    since   = datetime.utcnow() - timedelta(days=days)
    filters = "WHERE screened_at >= :since"
    params  = {"since": since}
    if risk_level:
        filters += " AND risk_level = :risk"
        params["risk"] = risk_level

    db = SessionLocal()
    try:
        rows = db.execute(text(f"""
            SELECT id, file_name, doc_type, applicant_id, id_number,
                   risk_level, risk_score, flags, screened_at,
                   full_result->'staff_decision'->>'decision' AS staff_decision,
                   full_result->'staff_decision'->>'notes' AS staff_notes
            FROM screening_logs {filters}
            ORDER BY screened_at DESC
        """), params).fetchall()

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "ID", "File Name", "Doc Type", "Applicant ID", "ID Number",
            "Risk Level", "Risk Score", "Flag Count", "Flags",
            "Screened At", "Staff Decision", "Staff Notes"
        ])
        for r in rows:
            flags = r.flags or []
            if isinstance(flags, str):
                try: flags = json.loads(flags)
                except: flags = []
            writer.writerow([
                r.id, r.file_name, r.doc_type, r.applicant_id or "",
                r.id_number or "", r.risk_level, r.risk_score,
                len(flags), " | ".join(flags),
                str(r.screened_at)[:19] if r.screened_at else "",
                r.staff_decision or "", r.staff_notes or "",
            ])
        return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    finally:
        db.close()


def export_cases_excel(days: int = 30, risk_level: str = "") -> bytes | None:
    """Export screening cases as Excel with formatting."""
    if not EXCEL_OK:
        return None

    since   = datetime.utcnow() - timedelta(days=days)
    filters = "WHERE screened_at >= :since"
    params  = {"since": since}
    if risk_level:
        filters += " AND risk_level = :risk"
        params["risk"] = risk_level

    db = SessionLocal()
    try:
        rows = db.execute(text(f"""
            SELECT id, file_name, doc_type, applicant_id, id_number,
                   risk_level, risk_score, flags, screened_at,
                   full_result->'staff_decision'->>'decision' AS staff_decision
            FROM screening_logs {filters}
            ORDER BY screened_at DESC
        """), params).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Screening Cases"

        # ── Header row ────────────────────────────────────────────────────────
        headers = ["ID", "File Name", "Doc Type", "Applicant ID", "ID Number",
                   "Risk Level", "Risk Score", "Flag Count", "Screened At", "Decision"]
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 25

        # ── Risk colors ───────────────────────────────────────────────────────
        risk_fills = {
            "HIGH":   PatternFill("solid", fgColor="FEE2E2"),
            "MEDIUM": PatternFill("solid", fgColor="FFFBEB"),
            "LOW":    PatternFill("solid", fgColor="F0FDF4"),
        }
        risk_fonts = {
            "HIGH":   Font(color="991B1B", bold=True, size=9),
            "MEDIUM": Font(color="92400E", bold=True, size=9),
            "LOW":    Font(color="166534", bold=True, size=9),
        }
        thin = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, r in enumerate(rows, 2):
            flags = r.flags or []
            if isinstance(flags, str):
                try: flags = json.loads(flags)
                except: flags = []

            level  = r.risk_level or "—"
            values = [
                r.id, r.file_name, (r.doc_type or "").upper(),
                r.applicant_id or "", r.id_number or "",
                level, r.risk_score, len(flags),
                str(r.screened_at)[:19] if r.screened_at else "",
                r.staff_decision or "PENDING",
            ]
            fill = risk_fills.get(level)
            font = risk_fonts.get(level, Font(size=9))

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin
                cell.font   = Font(size=9)
                cell.alignment = Alignment(vertical="center")
                if col == 6 and fill:  # risk level column
                    cell.fill = fill
                    cell.font = font

        # ── Column widths ─────────────────────────────────────────────────────
        widths = [6, 30, 12, 20, 20, 12, 10, 10, 20, 18]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        total  = len(rows)
        high   = sum(1 for r in rows if r.risk_level == "HIGH")
        medium = sum(1 for r in rows if r.risk_level == "MEDIUM")
        low    = sum(1 for r in rows if r.risk_level == "LOW")

        summary_data = [
            ["KYC Screening Summary Report"],
            [f"Period: Last {days} days"],
            [f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}"],
            [""],
            ["Metric", "Count", "Percentage"],
            ["Total Screened", total, "100%"],
            ["HIGH Risk",   high,   f"{high/max(total,1)*100:.1f}%"],
            ["MEDIUM Risk", medium, f"{medium/max(total,1)*100:.1f}%"],
            ["LOW Risk",    low,    f"{low/max(total,1)*100:.1f}%"],
        ]
        for row in summary_data:
            ws2.append(row)

        ws2["A1"].font = Font(bold=True, size=14, color="0F172A")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        db.close()


def export_monthly_stats_excel() -> bytes | None:
    """Export month-by-month statistics for the last 12 months."""
    if not EXCEL_OK:
        return None

    db = SessionLocal()
    try:
        rows = db.execute(text("""
            SELECT
                TO_CHAR(screened_at, 'YYYY-MM') AS month,
                COUNT(*) total,
                COUNT(*) FILTER (WHERE risk_level='HIGH')   high,
                COUNT(*) FILTER (WHERE risk_level='MEDIUM') medium,
                COUNT(*) FILTER (WHERE risk_level='LOW')    low,
                COUNT(*) FILTER (WHERE doc_type='id_card')  id_cards,
                COUNT(*) FILTER (WHERE doc_type='pdf')      pdfs,
                ROUND(AVG(risk_score), 1) avg_score
            FROM screening_logs
            WHERE screened_at >= NOW() - INTERVAL '12 months'
            GROUP BY TO_CHAR(screened_at, 'YYYY-MM')
            ORDER BY month DESC
        """)).fetchall()

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Monthly Stats"

        headers = ["Month", "Total", "HIGH", "MEDIUM", "LOW", "ID Cards", "PDFs", "Avg Score"]
        hfill   = PatternFill("solid", fgColor="0F172A")
        hfont   = Font(color="FFFFFF", bold=True, size=10)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(rows, 2):
            rd = dict(r._mapping)
            ws.append([rd["month"], rd["total"], rd["high"], rd["medium"],
                       rd["low"], rd["id_cards"], rd["pdfs"], float(rd["avg_score"] or 0)])

        for col in [1,2,3,4,5,6,7,8]:
            ws.column_dimensions[get_column_letter(col)].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        db.close()
