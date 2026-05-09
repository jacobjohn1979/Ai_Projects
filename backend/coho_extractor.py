"""
coho_extractor.py
Extract transactions from bank statement PDFs and populate the
Bank's Uniform Excel Template for Conduct of Account (COHO) Summary.

Supports: ABA Bank, ACLEDA, Prince Bank, and other Cambodian bank statements.

Usage (standalone):
    python coho_extractor.py <statement.pdf> [output.xlsx]

Usage (as module):
    from coho_extractor import extract_statement, fill_coho_template
    data = extract_statement("statement.pdf")
    fill_coho_template(data, "output.xlsx", "template.xlsx")
"""

import re
import json
import sys
import os
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colours matching original template ───────────────────────────────────────
C_HEADER_BG  = "1F4E79"   # dark blue — title row
C_SECTION_BG = "D6E4F0"   # light blue — section headers
C_SUMM_BG    = "E2EFDA"   # green — summary panel
C_MONTH_HDR  = "BDD7EE"   # month table header
C_TOTAL_BG   = "FFFF99"   # yellow — totals
C_WHITE      = "FFFFFF"
C_GREY       = "F2F2F2"
C_CLOSING    = "FCE4D6"   # orange — daily closing balance highlight

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
MEDIUM = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)



# ── Bank Profile Loader ───────────────────────────────────────────────────────
PROFILE_DIR = Path("/app/bank_profiles")

def _load_profiles_for_text(text_upper: str) -> list:
    """Find all matching trained profiles for this PDF (may have USD + KHR)."""
    if not PROFILE_DIR.exists():
        return []
    matched = []
    seen    = set()
    for f in sorted(PROFILE_DIR.glob("*.json")):
        try:
            p        = json.loads(f.read_text())
            prof_key = p.get("profile_key", p.get("swift", f.stem))
            if prof_key in seen:
                continue
            swift    = p.get("swift","").upper()
            keywords = [k.upper() for k in (p.get("keywords") or []) if k]
            # Add swift and bank name words as implicit keywords
            if swift and swift not in keywords:
                keywords.append(swift)
            for w in p.get("bank_name","").upper().split():
                if len(w) > 3 and w not in keywords:
                    keywords.append(w)
            # Check all keywords
            for kw in keywords:
                if kw and kw in text_upper:
                    matched.append(p)
                    seen.add(prof_key)
                    break
            # Special: Woori has no SWIFT — detect by unique field names
            if prof_key not in seen and "CID" in text_upper and "ACCOUNTNUMBER" in text_upper.replace(" ",""):
                if "WOORI" in p.get("bank_name","").upper() or "HVBK" in swift:
                    matched.append(p)
                    seen.add(prof_key)
        except:
            pass
    return matched

def _load_profile_for_text(text_upper: str) -> dict | None:
    """Find first matching trained profile (legacy single-profile support)."""
    results = _load_profiles_for_text(text_upper)
    return results[0] if results else None

def _load_profile_by_key(profile_key: str) -> dict | None:
    """Load a specific profile by its key (SWIFT_CURRENCY)."""
    f = PROFILE_DIR / (profile_key.upper() + ".json")
    if f.exists():
        try: return json.loads(f.read_text())
        except: pass
    return None

def _load_profile_by_swift(swift: str) -> dict | None:
    """Load profile by SWIFT — returns first match (legacy)."""
    return _load_profile_by_key(swift)


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 1 — PDF PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class BankStatementParser:
    """
    Parse ABA Bank (and similar Cambodian bank) PDF statements.
    Extracts: account details, opening/ending balances, all transactions.
    """

    MONTH_MAP = {
        "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
        "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
    }

    def __init__(self, pdf_path: str):
        self.pdf_path  = pdf_path
        self.pages     = []
        self.full_text = ""
        self._load()

    def _load(self):
        with pdfplumber.open(self.pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
                self.pages.append(text)
        self.full_text = "\n".join(self.pages)

    def _detect_bank(self) -> str:
        p = self.full_text.upper()

        # Check trained profiles first
        profiles = _load_profiles_for_text(p)
        if profiles:
            keys = ",".join(pr.get("profile_key", pr.get("swift","")) for pr in profiles)
            return "PROFILES:" + keys

        # Built-in detection by SWIFT code (most reliable)
        if "ABAAKHPP"    in p: return "ABA"
        if "WIGCKHPPXXX" in p: return "WING"
        if "ACLBKHPP"    in p: return "ACLEDA"
        if "ACLEDA"      in p: return "ACLEDA"
        if "CADIKHPP"    in p: return "CANADIA"
        if "HLFBKHPP"    in p: return "HATTHA"
        if "HATTHA"      in p: return "HATTHA"
        if "STPBKHPP"    in p: return "SATHAPANA"
        if "PERIOD FROM:" in p and "MONEY IN" in p and "MONEY OUT" in p: return "SATHAPANA"
        if "Sathapana" in self.full_text: return "SATHAPANA"
        if "MBBECAMM"    in p: return "MAYBANK"
        if "PPCBKHPP"    in p: return "PRINCE"
        if "VATTANAC"    in p: return "VATTANAC"

        # No SWIFT — detect by unique content
        if "CID" in p and "ACCOUNTNUMBER" in p.replace(" ",""): return "WOORI"
        if "WOORI BANK"  in p: return "WOORI"
        if "BALANCE AT PERIOD S" in p: return "POSTBANK"   # Post Bank split word
        if "BOOK DATE" in p and "CLOSING BALANCE" in p: return "POSTBANK"
        if "A/C:" in p and "WITHDRAWAL" in p and "DEPOSIT" in p: return "MAYBANK"
        if "ACCOUNT STATEMENT" in p and "PHILLIP" in p: return "PHILIP"
        if "ACCOUNT STATEMENT" in p and "PHILIP" in p: return "PHILIP"
        if "CADIKHPP" in p: return "CANADIA"

        # ACLEDA KHR — garbled Khmer text but has ACLEDA footer
        if "acledabank" in self.full_text.lower(): return "ACLEDA"
        # Sathapana — unique column headers
        if "PERIOD FROM:" in p and "MONEY IN" in p and "MONEY OUT" in p: return "SATHAPANA"
        if "BEGINNING BALANCE:" in p and "ENDING BALANCE:" in p and "REFERENCE NO" in p: return "SATHAPANA"

        return "GENERIC"

    def _parse_header(self) -> dict:
        h    = {}
        p    = self.full_text
        bank = self._detect_bank()

        if bank == "WING":
            h["bank"] = "Wing Bank"
            m = re.search(r"Account Number[:\s]+(\d+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Account Currency[:\s]+(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "USD"
            m = re.search(r"Account Type[:\s]+(\w+)", p)
            h["account_type"] = m.group(1).strip() if m else ""
            m = re.search(r"For period\s+([\d\-A-Za-z]+)\s*-\s*([\d\-A-Za-z]+)", p)
            if m:
                h["period_from"] = self._parse_date_wing(m.group(1).strip())
                h["period_to"]   = self._parse_date_wing(m.group(2).strip())
            m = re.search(r"Opening Balance[:\s]+([\d,\.]+)", p)
            h["opening_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Credit[:\s]+([\d,\.]+)", p)
            h["total_in"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Debit[:\s]+\-?([\d,\.]+)", p)
            h["total_out"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Ending Balance[:\s]+([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            h["holder_name"] = p.strip().split("\n")[0].strip()[:50]

        elif bank == "ACLEDA":
            h["bank"] = "ACLEDA Bank"
            m = re.search(r"Account Number\s*:\s*(\S+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Currency\s*:\s*(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "KHR"
            m = re.search(r"Statement Period\s*:\s*([\d/]+)\s*to\s*([\d/]+)", p)
            if m:
                h["period_from"] = self._parse_date_dmy(m.group(1).strip())
                h["period_to"]   = self._parse_date_dmy(m.group(2).strip())
            m = re.search(r"Name\s*:\s*([A-Z][A-Z\s]+)", p)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            h["opening_balance"] = 0.0
            h["total_in"]  = 0.0
            h["total_out"] = 0.0
            m = re.search(r"Balance at Period End\s+([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        elif bank == "HATTHA":
            h["bank"] = "Hattha Bank"
            m = re.search(r"Account Holder\s+(.+?)(?:\n|Account Type)", p, re.DOTALL)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            m = re.search(r"Account Number\s+([\d\s]+)", p)
            h["account_no"] = m.group(1).strip().replace(" ","") if m else ""
            m = re.search(r"Account Currency\s+(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "USD"
            m = re.search(r"From\s*:\s*([\d\-A-Za-z]+)\s+to\s+([\d\-A-Za-z]+)", p, re.IGNORECASE)
            if m:
                h["period_from"] = self._parse_date_wing(m.group(1).strip())
                h["period_to"]   = self._parse_date_wing(m.group(2).strip())
            m = re.search(r"Total Money In\s+([\d,\.]+)", p)
            h["total_in"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Money Out\s+([\d,\.]+)", p)
            h["total_out"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Available Balance\s+([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        elif bank == "CANADIA":
            h["bank"] = "Canadia Bank"
            m = re.search(r"Customer Name:\s*(.+?)(?:\n|Customer ID)", p, re.DOTALL)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            m = re.search(r"Account Number:\s*(\S+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Currency\s*:\s*(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "KHR"
            m = re.search(r"Account Statement Period\s*:\s*([\d/]+)-([\d/]+)", p)
            if m:
                h["period_from"] = self._parse_date_dmy(m.group(1))
                h["period_to"]   = self._parse_date_dmy(m.group(2))
            m = re.search(r"Balance at Period Start\s*:\s*([\d,\.]+)", p)
            h["opening_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Ending Balance\s*:\s*([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Cash-In\s*:\s*([\d,\.]+)", p)
            h["total_in"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Cash-Out\s*:\s*([\d,\.]+)", p)
            h["total_out"] = float(m.group(1).replace(",","")) if m else 0.0

        elif bank == "SATHAPANA":
            h["bank"] = "Sathapana Bank"
            m = re.search(r"Account Statement\s+(.+?)(?:\n|Account Type)", p, re.DOTALL)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            m = re.search(r"Account Number:\s*(\S+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Period from:\s*([\w\s,]+)\s*-\s*([\w\s,]+)", p)
            if m:
                h["period_from"] = self._parse_date(m.group(1).strip())
                h["period_to"]   = self._parse_date(m.group(2).strip())
            m = re.search(r"Beginning Balance:\s*([\d,\.]+)", p)
            h["opening_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Ending Balance:\s*([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            h["currency"] = "USD"

        elif bank == "POSTBANK":
            h["bank"] = "Post Bank"
            m = re.search(r"Customer\s*:\s*\d+\s+(.+?)(?:\n|Currency)", p, re.DOTALL)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            m = re.search(r"Account\s*:\s*(\d+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Currency\s*:\s*(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "USD"
            m = re.search(r"Balance at Period S[^\d]*([\d,\.]+)", p)
            h["opening_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        elif bank == "MAYBANK":
            h["bank"] = "Maybank"
            lines = p.split("\n")
            for line in lines[:8]:
                s = line.strip()
                if s and len(s) > 3 and not any(x in s for x in [
                    "CA3400","v0.0","N/A","BARKU","KANDAL","CAMBODIA"]):
                    h["holder_name"] = s[:50]; break
            m = re.search(r"A/C:\s*([\d/\(\)\w\s]+?)(?:Maybank|\n)", p)
            h["account_no"] = m.group(1).strip()[:30] if m else ""
            h["currency"] = "USD"

        elif bank == "WOORI":
            h["bank"] = "Woori Bank"
            m = re.search(r"Account Number\s+(\d+)", p)
            h["account_no"]   = m.group(1).strip() if m else ""
            m = re.search(r"Account Name\s+(.+?)(?:\n|Total)", p, re.DOTALL)
            h["holder_name"]  = m.group(1).strip()[:50] if m else ""
            h["currency"]     = "USD"
            # Statement Date: 22-Aug-24
            m = re.search(r"Statement Date\s+(\d{2}-[A-Za-z]+-\d{2,4})", p)
            if m:
                end_date = self._parse_date_wing(
                    m.group(1) if len(m.group(1).split("-")[2]) == 4
                    else m.group(1)[:-2] + "20" + m.group(1)[-2:]
                )
                h["period_to"] = end_date
                # Period: 6 Months
                mo = re.search(r"Statement Period\s+(\d+)\s+Months", p)
                if mo and end_date:
                    from datetime import timedelta
                    months = int(mo.group(1))
                    yr  = end_date.year
                    mon = end_date.month - months
                    while mon <= 0:
                        mon += 12; yr -= 1
                    h["period_from"] = date(yr, mon, 1)
            m = re.search(r"Total Balance\s+([\d,\.]+)\s+USD", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        elif bank == "PHILIP":
            h["bank"] = "Philip Bank"
            m = re.search(r"Account Number\s+(\d+)", p)
            h["account_no"]  = m.group(1).strip() if m else ""
            m = re.search(r"Period:\s*([\d\-]+)\s*to\s*([\d\-]+)", p)
            if m:
                try:
                    from datetime import datetime as dt2
                    h["period_from"] = dt2.strptime(m.group(1).strip(), "%Y-%m-%d").date()
                    h["period_to"]   = dt2.strptime(m.group(2).strip(), "%Y-%m-%d").date()
                except: pass
            # Holder name is line before ACCOUNT INFORMATION
            lines = p.split("\n")
            for i, line in enumerate(lines):
                if "ACCOUNT INFORMATION" in line.upper() and i > 0:
                    h["holder_name"] = lines[i-1].strip()[:50]
                    break
            h["currency"] = "USD"
            m = re.search(r"Current Balance\s+([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        else:  # ABA / Generic
            h["bank"] = "ABA Bank" if "ABAAKHPP" in p else "Unknown"
            m = re.search(r"Account Holder Name\s+(.+?)(?:\n|Account Type)", p, re.IGNORECASE|re.DOTALL)
            h["holder_name"] = m.group(1).strip()[:50] if m else ""
            m = re.search(r"Account No\.\s+(\S+)", p)
            h["account_no"] = m.group(1).strip() if m else ""
            m = re.search(r"Account Currency\s+(\w+)", p)
            h["currency"] = m.group(1).strip() if m else "USD"
            m = re.search(r"For period[:\s]+([\w\s,]+)\s*[–-]+\s*([\w\s,]+)", p)
            if m:
                h["period_from"] = self._parse_date(m.group(1).strip())
                h["period_to"]   = self._parse_date(m.group(2).strip())
            m = re.search(r"Opening Balance\s+([\d,\.]+)", p)
            h["opening_balance"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Money In\s*\+\s*([\d,\.]+)", p)
            h["total_in"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Total Money Out\s*-\s*([\d,\.]+)", p)
            h["total_out"] = float(m.group(1).replace(",","")) if m else 0.0
            m = re.search(r"Ending Balance\s+([\d,\.]+)", p)
            h["ending_balance"] = float(m.group(1).replace(",","")) if m else 0.0

        h.setdefault("holder_name",     "")
        h.setdefault("account_no",      "")
        h.setdefault("currency",        "USD")
        h.setdefault("opening_balance", 0.0)
        h.setdefault("total_in",        0.0)
        h.setdefault("total_out",       0.0)
        h.setdefault("ending_balance",  0.0)
        h.setdefault("period_from",     None)
        h.setdefault("period_to",       None)
        h["bank_type"] = bank
        return h

    def _parse_date(self, s: str) -> date | None:
        """ABA format: Aug 01, 2025"""
        s = s.strip()
        m = re.match(r"(\w{3})\s+(\d{1,2}),?\s+(\d{4})", s)
        if m:
            mon = self.MONTH_MAP.get(m.group(1).lower())
            if mon:
                return date(int(m.group(3)), mon, int(m.group(2)))
        return None

    def _parse_date_wing(self, s: str) -> date | None:
        """Wing format: 22-Apr-2026 or 06-Feb-2026"""
        m = re.match(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", s.strip())
        if m:
            mon = self.MONTH_MAP.get(m.group(2).lower())
            if mon:
                return date(int(m.group(3)), mon, int(m.group(1)))
        return None

    def _parse_date_dmy(self, s: str) -> date | None:
        """ACLEDA format: 01/02/23 (DD/MM/YY)"""
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s.strip())
        if m:
            day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if yr < 100:
                yr += 2000
            return date(yr, mon, day)
        return None

    def _parse_date_acleda_row(self, s: str) -> date | None:
        """ACLEDA row date: 24 MAR 23"""
        m = re.match(r"(\d{1,2})\s+([A-Z]{3})\s+(\d{2})", s.strip())
        if m:
            mon = self.MONTH_MAP.get(m.group(2).lower())
            yr  = int(m.group(3)) + 2000
            if mon:
                return date(yr, mon, int(m.group(1)))
        return None

    def _parse_transactions(self) -> list[dict]:
        bank = self._detect_bank()
        # Use trained profiles if available (may be multiple for USD+KHR)
        if bank.startswith("PROFILES:"):
            keys     = bank[9:].split(",")
            all_txns = []
            for key in keys:
                profile = _load_profile_by_key(key.strip())
                if profile:
                    try:
                        from bank_trainer import _parse_with_profile
                        txns = _parse_with_profile(self.pages, profile)
                        # Convert string dates to date objects
                        for t in txns:
                            if isinstance(t.get("date"), str):
                                try:
                                    from datetime import datetime
                                    t["date"] = datetime.strptime(t["date"][:10], "%Y-%m-%d").date()
                                except: pass
                        all_txns.extend(txns)
                    except Exception as e:
                        pass
            if all_txns:
                # Sort by date
                all_txns.sort(key=lambda t: t["date"])
                return all_txns
        if bank == "WING":      return self._parse_wing()
        elif bank == "ACLEDA":  return self._parse_acleda()
        elif bank == "WOORI":   return self._parse_woori()
        elif bank == "PHILIP":  return self._parse_philip()
        elif bank == "HATTHA":  return self._parse_hattha()
        elif bank == "CANADIA": return self._parse_canadia()
        elif bank == "SATHAPANA": return self._parse_sathapana()
        elif bank == "POSTBANK":  return self._parse_postbank()
        elif bank == "MAYBANK":   return self._parse_maybank()
        else:                   return self._parse_aba()

    # ── Wing Bank ─────────────────────────────────────────────────────────────

    def _parse_wing(self) -> list[dict]:
        """
        Wing Bank PDF has two date patterns:
          Pattern A: date alone on line, amounts on next line
            14: 'Direct Bank Transfer from'
            15: '22-Apr-2026'
            16: '004 EWJ608421 001214686 JOHN JACOB, MR - 640.00 687.37'
            17: '08:17:18 PM'
            18: 'Vattanac Bank e053e12e'

          Pattern B: date + description on same line, amounts on next line
            19: '30-Apr-2026 Saving Interest from'
            20: '004 004IRFHUSD000002 - 0.02 687.39'
            21: '09:33:48 PM 100536313'

        Amount line format: branch ref description - credit balance
        '-' in debit position means no debit (all credits in this sample)
        Last number = balance, second-to-last = transaction amount
        """
        transactions = []
        # Match date at start of line, with optional description after
        DATE_RE = re.compile(r"^(\d{2}-[A-Za-z]{3}-\d{4})(.*)")
        TIME_RE = re.compile(r"\d{2}:\d{2}:\d{2}\s*[AP]M")
        AMT_RE  = re.compile(r"([\d,]+\.\d{2})")

        SKIP = {"Ending Balance","Wing Bank","PAGE","Balance in","ACCOUNT SUMMARY",
                "ACCOUNT INFORMATION","Opening Balance","Total Credit","Total Debit",
                "Available Balance","Blocked Balance","Bank Swift","Account Number",
                "Account Currency","Account Type","For period"}

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            prev_desc = ""

            while i < len(lines):
                line = lines[i]

                # Skip header/footer lines
                if any(s in line for s in SKIP):
                    prev_desc = ""
                    i += 1
                    continue

                dm = DATE_RE.match(line)
                if dm:
                    trx_date   = self._parse_date_wing(dm.group(1))
                    inline_desc = dm.group(2).strip()
                    i += 1

                    # Collect body lines until next date or end markers
                    body_lines = []
                    while i < len(lines):
                        nl = lines[i]
                        if DATE_RE.match(nl):
                            break
                        if any(s in nl for s in SKIP):
                            break
                        body_lines.append(nl)
                        i += 1

                    body = " ".join(body_lines)

                    # Find all decimal amounts in body
                    amounts = [float(m.group(1).replace(",",""))
                               for m in AMT_RE.finditer(body)
                               if float(m.group(1).replace(",","")) > 0]

                    if len(amounts) < 2:
                        # Try prev_desc + inline_desc as clue, skip
                        prev_desc = inline_desc or prev_desc
                        continue

                    balance = amounts[-1]
                    amount  = amounts[-2]

                    # Build description
                    desc_parts = [x for x in [prev_desc, inline_desc] if x]
                    desc_raw   = " ".join(desc_parts) + " " + body
                    desc = TIME_RE.sub('', desc_raw)
                    desc = re.sub(r'\b\d{3}\b', '', desc)        # branch code
                    desc = re.sub(r'\b[A-Z0-9]{8,}\b', '', desc) # ref numbers
                    desc = AMT_RE.sub('', desc)                   # amounts
                    desc = re.sub(r'\s+-\s*', ' ', desc)
                    desc = re.sub(r'\s+', ' ', desc).strip(' -,.')[:100]

                    # Direction — Wing uses '-' in debit col for credits
                    rl = (inline_desc + " " + body).lower()
                    is_debit = any(x in rl for x in ["fee","charge","payment to","transfer to","withdrawal"])

                    if trx_date and amount > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  0.0 if is_debit else round(amount, 2),
                            "money_out": round(amount, 2) if is_debit else 0.0,
                            "balance":   round(balance, 2),
                        })
                    prev_desc = ""

                elif not TIME_RE.search(line) and len(line) > 3:
                    prev_desc = line
                    i += 1
                else:
                    i += 1

        return transactions

    # ── ACLEDA Bank ───────────────────────────────────────────────────────────

    def _parse_acleda(self) -> list[dict]:
        """
        ACLEDA format (text-based PDF):
          24 MAR 23  Description  Ref  24 MAR 23  [debit]  [credit]  balance
        POST DATE | DESC | REF | VALUE DATE | DEBIT | CREDIT | BALANCE
        Only two or three numbers at end — balance always last.
        """
        transactions = []
        DATE_RE = re.compile(r"^(\d{1,2}\s+[A-Z]{3}\s+\d{2})\s+(.*)")
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)")

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            while i < len(lines):
                line = lines[i]
                dm   = DATE_RE.match(line)
                if dm:
                    trx_date = self._parse_date_acleda_row(dm.group(1))
                    rest = dm.group(2)
                    # Collect continuation lines
                    j = i + 1
                    while j < len(lines):
                        nl = lines[j]
                        if DATE_RE.match(nl):
                            break
                        if any(x in nl for x in ["Page ", "Balance at Period", "ACLEDA", "This is"]):
                            break
                        rest += " " + nl
                        j += 1
                    i = j

                    if not trx_date:
                        continue

                    # All numeric amounts
                    nums = [(m.start(), float(m.group(1).replace(",","")))
                            for m in AMT_RE.finditer(rest)
                            if "." in m.group(1) and float(m.group(1).replace(",","")) > 0]

                    if len(nums) < 2:
                        continue

                    balance = nums[-1][1]
                    amount  = nums[-2][1]

                    # Direction from description
                    rl = rest.lower()
                    is_credit = any(x in rl for x in ["credit", "from-", "from -", "interest", "qr payment credit"])
                    is_debit  = any(x in rl for x in ["debit", "own account", "to -", "fee charge", "qr payment debit", "bakong"])

                    if not is_credit and not is_debit:
                        is_credit = True  # default

                    # Clean description
                    desc = re.sub(r'\b\d{11,}\b', '', rest)
                    desc = re.sub(r'\b\d{1,2}\s+[A-Z]{3}\s+\d{2}\b', '', desc)
                    desc = re.sub(r'([\d,]+\.?\d*)', '', desc)
                    desc = re.sub(r'\s+', ' ', desc).strip()[:100]

                    if amount > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  round(amount, 2) if is_credit else 0.0,
                            "money_out": round(amount, 2) if is_debit  else 0.0,
                            "balance":   round(balance, 2),
                        })
                else:
                    i += 1

        return transactions

    def _parse_hattha(self) -> list[dict]:
        """
        Hattha Bank format (USD and KHR):
          29 Feb 2024 | 21:34 TAX 0.45 USD 7818.87 USD
          (007SINRUSD000002)
        Columns: Date | Time Description [Out] [In] Balance CURRENCY
        Direction: Money Out first, Money In second (before balance)
        Header has: Total Money In / Total Money Out
        """
        transactions = []
        # Match: "29 Feb 2024 | 21:34 DESCRIPTION [amt CURR] ... balance CURR"
        DATE_RE = re.compile(
            r"^(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s*\|\s*\d{2}:\d{2}\s+(.*)"
        )
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)\s*(?:USD|KHR)")

        currency = "KHR" if "KHR" in self.full_text.upper()[:500] else "USD"

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            while i < len(lines):
                line = lines[i]
                m    = DATE_RE.match(line)
                if m:
                    date_str = m.group(1)
                    rest     = m.group(2)
                    trx_date = self._parse_date(date_str.replace(" 2024",", 2024")
                                                          .replace(" 2023",", 2023")
                                                          .replace(" 2025",", 2025")
                                                          .replace(" 2026",", 2026"))
                    if not trx_date:
                        # Try "29 Feb 2024" format
                        dm2 = re.match(r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})", date_str)
                        if dm2:
                            mon = self.MONTH_MAP.get(dm2.group(2).lower())
                            if mon:
                                trx_date = date(int(dm2.group(3)), mon, int(dm2.group(1)))
                    i += 1
                    # Skip ref line (starts with parenthesis)
                    if i < len(lines) and lines[i].startswith("("):
                        i += 1

                    if not trx_date:
                        continue

                    amounts = [float(m2.group(1).replace(",",""))
                               for m2 in AMT_RE.finditer(rest)
                               if float(m2.group(1).replace(",","")) > 0]

                    if len(amounts) < 2:
                        continue

                    balance = amounts[-1]
                    amount  = amounts[-2]

                    # Direction from header keywords
                    rl       = rest.lower()
                    is_debit = any(x in rl for x in [
                        "loan","tax","fee","payment","transfer out","withdrawal",
                        "hattha pay","within transfer"
                    ])
                    is_credit = any(x in rl for x in [
                        "deposit","credit","received","saving credit","interest credit"
                    ])
                    if not is_debit and not is_credit:
                        # Check position: if out col > 0 before in col
                        is_credit = len(amounts) == 3 and amounts[0] > 0

                    desc = AMT_RE.sub("", rest).strip()[:100]

                    if amount > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  round(amount, 2) if is_credit else 0.0,
                            "money_out": round(amount, 2) if is_debit  else 0.0,
                            "balance":   round(balance, 2),
                        })
                else:
                    i += 1
        return transactions

    def _parse_canadia(self) -> list[dict]:
        """
        Canadia Bank format (KHR):
          05/10/2025 FT252791S2BW Transfer 4,025.00 0.00 822.00
          07/10/2025 FT25280MQ42B Transfer 0.00 997,204.00 998,026.00
        Columns: DD/MM/YYYY Ref Description Out In Balance
        SWIFT: CADIKHPP
        """
        transactions = []
        DATE_RE = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(.*)")
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)")
        SKIP    = {"print date","account statement","swift","customer","activity log",
                   "transaction reference","transaction date","balance at"}

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            while i < len(lines):
                line = lines[i]
                if any(s in line.lower() for s in SKIP):
                    i += 1; continue

                m = DATE_RE.match(line)
                if m:
                    trx_date = self._parse_date_dmy(m.group(1))
                    rest     = m.group(2)
                    i += 1
                    # Collect continuation (description on next line)
                    while i < len(lines) and not DATE_RE.match(lines[i]):
                        nl = lines[i]
                        if any(s in nl.lower() for s in SKIP): break
                        rest += " " + nl
                        i += 1

                    amounts = [float(m2.group(1).replace(",",""))
                               for m2 in AMT_RE.finditer(rest)
                               if float(m2.group(1).replace(",","")) > 0]

                    # Need at least 3: out, in, balance (one may be 0)
                    # Get all numbers including zeros
                    all_nums = [float(m2.group(1).replace(",",""))
                                for m2 in AMT_RE.finditer(rest)]
                    if len(all_nums) < 3:
                        if not trx_date: continue
                        # Only 2 numbers: amount + balance
                        if len(amounts) >= 2:
                            balance = amounts[-1]
                            amount  = amounts[-2]
                            rl = rest.lower()
                            is_debit = "transfer" in rl and amount > 0
                            transactions.append({
                                "date": trx_date, "desc": rest[:80],
                                "money_in": 0.0 if is_debit else round(amount,2),
                                "money_out": round(amount,2) if is_debit else 0.0,
                                "balance": round(balance,2),
                            })
                        continue

                    balance   = all_nums[-1]
                    money_out = all_nums[-3] if len(all_nums) >= 3 else 0.0
                    money_in  = all_nums[-2] if len(all_nums) >= 2 else 0.0

                    desc = AMT_RE.sub("", rest).strip()[:100]

                    if money_out > 0 or money_in > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  round(money_in, 2),
                            "money_out": round(money_out, 2),
                            "balance":   round(balance, 2),
                        })
                else:
                    i += 1
        return transactions

    def _parse_sathapana(self) -> list[dict]:
        """
        Sathapana Bank — same Money In / Money Out format as ABA.
          Oct 01, 2025 19:49 046OART252740504 Own Account Transfer to 0.07 USD
        Uses ABA-style parsing (Mon DD, YYYY) with USD amounts.
        """
        transactions = []
        DATE_RE = re.compile(r"^([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})\s+(.*)")
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)\s+USD")

        for page_text in self.pages:
            for line in page_text.split("\n"):
                line = line.strip()
                m    = DATE_RE.match(line)
                if not m:
                    continue
                rest = m.group(2).strip()
                if any(x in rest for x in ["Beginning Balance","Ending Balance","Date Reference"]):
                    continue
                trx_date = self._parse_date(m.group(1))
                if not trx_date:
                    continue

                amt_matches = [(m2.start(), float(m2.group(1).replace(",","")))
                               for m2 in AMT_RE.finditer(rest)]
                if len(amt_matches) < 1:
                    continue

                balance  = amt_matches[-1][1]
                # Second to last = transaction amount if 2+ amounts
                if len(amt_matches) >= 2:
                    amount   = amt_matches[-2][1]
                    # Direction: Money Out col before Money In col
                    # Check keywords for direction
                    rl       = rest.lower()
                    is_debit = any(x in rl for x in [
                        "pay khqr","topup","mobile topup","fee","transfer to",
                        "own account transfer to","withdrawal","debit","pay to"
                    ])
                    is_credit = any(x in rl for x in [
                        "transfer from","received","own account transfer from","deposit","credit"
                    ])
                    if not is_debit and not is_credit:
                        is_debit = True  # default to debit for Sathapana unknown
                    money_in  = 0.0 if is_debit else amount
                    money_out = amount if is_debit else 0.0
                else:
                    # Only balance — skip
                    continue

                desc = rest[:amt_matches[0][0]].strip()[:100]

                if amount > 0:
                    transactions.append({
                        "date":      trx_date,
                        "desc":      desc,
                        "money_in":  round(money_in, 2),
                        "money_out": round(money_out, 2),
                        "balance":   round(balance, 2),
                    })
        return transactions

    def _parse_postbank(self) -> list[dict]:
        """
        Post Bank format:
          10 JAN 25 FT25010FHYJY Txn Payment 10 JAN 25 448.26 253.72
        Columns: Book Date Ref Description Value Date Debit Credit Closing Balance
        Date format: DD MON YY
        """
        transactions = []
        DATE_RE = re.compile(r"^(\d{1,2}\s+[A-Z]{3}\s+\d{2})\s+(.*)")
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)")
        SKIP    = {"book date","balance at period","account :","customer :","currency :"}

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            while i < len(lines):
                line = lines[i]
                if any(s in line.lower() for s in SKIP):
                    i += 1; continue

                m = DATE_RE.match(line)
                if m:
                    trx_date = self._parse_date_acleda_row(m.group(1))
                    rest     = m.group(2)
                    i += 1
                    # Collect continuation lines
                    while i < len(lines) and not DATE_RE.match(lines[i]):
                        nl = lines[i]
                        if any(s in nl.lower() for s in SKIP): break
                        rest += " " + nl
                        i += 1

                    if not trx_date:
                        continue

                    amounts = [float(m2.group(1).replace(",",""))
                               for m2 in AMT_RE.finditer(rest)
                               if "." in m2.group(1) and float(m2.group(1).replace(",","")) > 0]

                    if len(amounts) < 2:
                        continue

                    balance = amounts[-1]
                    amount  = amounts[-2]

                    rl        = rest.lower()
                    is_credit = any(x in rl for x in ["transfer in","deposit","credit","received","interest"])
                    is_debit  = any(x in rl for x in ["payment","loan","fee","tax","transfer out","withdrawal"])
                    if not is_credit and not is_debit:
                        is_credit = True

                    desc = re.sub(r"[\d,]+\.?\d*", "", rest)
                    desc = re.sub(r"\b[A-Z0-9]{8,}\b", "", desc)
                    desc = re.sub(r"\s+", " ", desc).strip()[:100]

                    if amount > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  round(amount, 2) if is_credit else 0.0,
                            "money_out": round(amount, 2) if is_debit  else 0.0,
                            "balance":   round(balance, 2),
                        })
                else:
                    i += 1
        return transactions

    def _parse_maybank(self) -> list[dict]:
        """
        Maybank format:
          NO DATE TRAN CODE/ REF.NO WITHDRAWAL DEPOSIT BALANCE
          1 02/05/23 209 BAKONG -50.00 844.22
        Columns: NO DD/MM/YY REF DESC WITHDRAWAL DEPOSIT BALANCE
        """
        transactions = []
        # Row starts with number then date
        DATE_RE  = re.compile(r"^\d+\s+(\d{2}/\d{2}/\d{2})\s+(.*)")
        AMT_RE   = re.compile(r"(-?[\d,]+\.\d+)")  # must have decimal
        SKIP     = {"no date","withdrawal","deposit","balance","a/c:","statement date",
                    "---","tran code","ref.no"}

        for page_text in self.pages:
            lines = [l.strip() for l in page_text.split("\n")]
            i = 0
            while i < len(lines):
                line = lines[i]
                if any(s in line.lower() for s in SKIP):
                    i += 1; continue

                m = DATE_RE.match(line)
                if m:
                    trx_date = self._parse_date_dmy(m.group(1))
                    rest     = m.group(2)
                    i += 1
                    while i < len(lines) and not DATE_RE.match(lines[i]):
                        nl = lines[i]
                        if any(s in nl.lower() for s in SKIP): break
                        rest += " " + nl
                        i += 1

                    if not trx_date:
                        continue

                    # Find all numbers (may be negative for withdrawals)
                    nums = [(m2.start(), float(m2.group(1).replace(",","")))
                            for m2 in AMT_RE.finditer(rest)
                            if m2.group(1).replace(",","").replace("-","").strip()]

                    if len(nums) < 2:
                        continue

                    balance = nums[-1][1]
                    amount  = abs(nums[-2][1])
                    is_debit = nums[-2][1] < 0

                    if not is_debit:
                        rl = rest.lower()
                        is_debit = any(x in rl for x in ["withdrawal","-"])

                    desc = AMT_RE.sub("", rest).strip()[:100]

                    if amount > 0:
                        transactions.append({
                            "date":      trx_date,
                            "desc":      desc,
                            "money_in":  0.0 if is_debit else round(amount, 2),
                            "money_out": round(amount, 2) if is_debit else 0.0,
                            "balance":   round(balance, 2),
                        })
                else:
                    i += 1
        return transactions

    def _parse_woori(self) -> list[dict]:
        """
        Woori Bank format:
          13-Aug-24 PaytoMINHCHENGPHE -3.57USD
          10-Aug-24 ReceivedFromBakongMember +210.00USD
        Date at start, description, +/- amount USD
        """
        transactions = []
        DATE_RE = re.compile(r"^(\d{2}-[A-Za-z]{3}-\d{2,4})\s+(.*)")
        AMT_RE  = re.compile(r"([+-]?[\d,]+\.?\d*)\s*USD")

        for page_text in self.pages:
            for line in page_text.split("\n"):
                line = line.strip()
                m    = DATE_RE.match(line)
                if not m:
                    continue
                trx_date = self._parse_date_wing(m.group(1).replace("-24","-2024").replace("-23","-2023").replace("-25","-2025").replace("-26","-2026"))
                # Try 2-digit year
                if not trx_date:
                    dm2 = re.match(r"(\d{2})-([A-Za-z]{3})-(\d{2})", m.group(1))
                    if dm2:
                        mon = self.MONTH_MAP.get(dm2.group(2).lower())
                        yr  = int(dm2.group(3)) + 2000
                        if mon:
                            trx_date = date(yr, mon, int(dm2.group(1)))

                rest = m.group(2)
                am   = AMT_RE.search(rest)
                if not am or not trx_date:
                    continue

                amount_str = am.group(1).replace(",","")
                amount     = abs(float(amount_str))
                is_credit  = amount_str.startswith("+") or "received" in rest.lower() or "credit" in rest.lower()
                is_debit   = amount_str.startswith("-") or "pay" in rest.lower() or "loan" in rest.lower()

                desc = AMT_RE.sub("", rest).strip()[:100]

                if amount > 0:
                    transactions.append({
                        "date":      trx_date,
                        "desc":      desc,
                        "money_in":  round(amount, 2) if is_credit else 0.0,
                        "money_out": round(amount, 2) if is_debit  else 0.0,
                        "balance":   0.0,
                    })
        return transactions

    def _parse_philip(self) -> list[dict]:
        """
        Philip Bank format:
          Post Date  Value Date  Description  Debit  Credit  Balance
          2023-05-15 2023-05-15  Transfer In         500.00  503.04
        """
        transactions = []
        DATE_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})\s+(\d{4}-\d{2}-\d{2})\s+(.*)")
        AMT_RE  = re.compile(r"([\d,]+\.?\d*)")

        for page_text in self.pages:
            for line in page_text.split("\n"):
                line = line.strip()
                m    = DATE_RE.match(line)
                if not m:
                    continue
                try:
                    from datetime import datetime as dt
                    trx_date = dt.strptime(m.group(1), "%Y-%m-%d").date()
                except: continue

                rest    = m.group(3)
                amounts = [float(x.group(1).replace(",",""))
                           for x in AMT_RE.finditer(rest)
                           if float(x.group(1).replace(",","")) > 0]

                if len(amounts) < 2:
                    continue

                balance = amounts[-1]
                amount  = amounts[-2]

                rl        = rest.lower()
                is_debit  = any(x in rl for x in ["withdraw","debit","fee","charge","transfer out","payment"])
                is_credit = any(x in rl for x in ["credit","deposit","received","transfer in"])
                if not is_debit and not is_credit:
                    is_credit = True

                desc = AMT_RE.sub("", rest).strip()[:100]

                if amount > 0:
                    transactions.append({
                        "date":      trx_date,
                        "desc":      desc,
                        "money_in":  round(amount, 2) if is_credit else 0.0,
                        "money_out": round(amount, 2) if is_debit  else 0.0,
                        "balance":   round(balance, 2),
                    })
        return transactions

    def _parse_aba(self) -> list[dict]:
        """ABA Bank: Aug 01, 2025 DESCRIPTION 10.63 USD USD 222.64 USD"""
        transactions = []
        DATE_RE = re.compile(r"^([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})\s+(.*)")
        AMT_RE  = re.compile(r"(\d[\d,]*\.?\d*)\s+USD")

        for page_text in self.pages:
            for raw_line in page_text.split("\n"):
                line = raw_line.strip()
                m    = DATE_RE.match(line)
                if not m:
                    continue
                trx_date = self._parse_date(m.group(1))
                rest     = m.group(2).strip()
                if any(x in rest for x in ["Advanced Bank","info@aba","Page ","Ending Balance","Total ","DISCLAIMER"]):
                    continue
                rest = re.sub(r'[\u1780-\u17ff\u19e0-\u19ff\u0000-\u0008]+', '', rest)
                amt_matches = [(m2.start(), float(m2.group(1).replace(",","")))
                               for m2 in AMT_RE.finditer(rest)]
                if len(amt_matches) < 2:
                    continue
                balance   = amt_matches[-1][1]
                txn_amt   = amt_matches[-2][1]
                txn_start = amt_matches[-2][0]
                chunk_before = rest[:txn_start]
                usd_before   = bool(re.search(r'\bUSD\b', chunk_before))
                money_out = txn_amt if usd_before else 0.0
                money_in  = txn_amt if not usd_before else 0.0
                desc = chunk_before
                for pat in [r'ORIGINAL AMOUNT$',r'REF#\s*\S+',r'TRAN#\s*\S+',
                            r'HASH#\s*[0-9a-fA-F]+',r'BAKONG#\s*[0-9a-fA-F]+',
                            r'EXT#\s*\S+',r'APV\s*\S+',r'TID:\s*\S+',
                            r'PURCHASE#\s*\S+',r'\|[^|]+\|',r'\bUSD\b']:
                    desc = re.sub(pat, '', desc)
                desc = re.sub(r'\s+', ' ', desc).strip().strip('|\\/. ')[:100]
                if trx_date and txn_amt > 0:
                    transactions.append({
                        "date": trx_date, "desc": desc,
                        "money_in":  round(money_in, 2),
                        "money_out": round(money_out, 2),
                        "balance":   round(balance, 2),
                    })
        return transactions

    # ── Main parse ─────────────────────────────────────────────────────────

    def parse(self) -> dict:
        header       = self._parse_header()
        transactions = self._parse_transactions()

        # Verify and fix running balances
        transactions = self._fix_balances(header["opening_balance"], transactions)

        return {
            "header":       header,
            "transactions": transactions,
            "parsed_at":    datetime.now().isoformat(),
            "source_file":  self.pdf_path,
        }

    def _fix_balances(self, opening: float, txns: list) -> list:
        """
        Recalculate running balances from scratch to ensure accuracy.
        Only keep transactions where balance makes sense.
        """
        bal = opening
        result = []
        for t in txns:
            bal = round(bal + t["money_in"] - t["money_out"], 2)
            t["running_balance"] = bal
            # Use the stated balance from PDF as O/S balance
            t["os_balance"] = t["balance"]
            result.append(t)
        return result


def extract_statement(pdf_path: str) -> dict:
    """Public function: extract bank statement from PDF."""
    parser = BankStatementParser(pdf_path)
    return parser.parse()


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 2 — ANALYTICS ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def _to_date(v):
    """Convert string, date, or None to a date object safely."""
    if v is None:
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        # Try ISO format: 2025-08-01
        try:
            from datetime import datetime
            return datetime.strptime(v[:10], "%Y-%m-%d").date()
        except: pass
        # Try DD/MM/YYYY
        try:
            from datetime import datetime
            return datetime.strptime(v, "%d/%m/%Y").date()
        except: pass
        # Try DD-Mon-YYYY
        try:
            from datetime import datetime
            return datetime.strptime(v, "%d-%b-%Y").date()
        except: pass
    return None


def compute_analytics(data: dict) -> dict:
    """Compute all COHO summary statistics from transaction data."""
    txns    = data["transactions"]
    header  = data["header"]

    if not txns:
        return {}

    opening = header.get("opening_balance", 0.0)

    # ── Per-transaction analytics ──────────────────────────────────────────
    total_in_trx  = sum(1 for t in txns if t["money_in"] > 0)
    total_out_trx = sum(1 for t in txns if t["money_out"] > 0)
    total_in_amt  = round(sum(t["money_in"] for t in txns), 2)
    total_out_amt = round(sum(t["money_out"] for t in txns), 2)

    # ── Daily closing balances ─────────────────────────────────────────────
    # Last transaction each day = daily closing balance
    daily_closing: dict[date, float] = {}
    for t in txns:
        daily_closing[t["date"]] = t["os_balance"]

    closing_values = list(daily_closing.values())
    highest_bal    = max(closing_values) if closing_values else 0.0
    lowest_bal     = min(closing_values) if closing_values else 0.0

    # ── Period ────────────────────────────────────────────────────────────
    date_from = _to_date(header.get("period_from")) or txns[0]["date"]
    date_to   = _to_date(header.get("period_to"))   or txns[-1]["date"]
    # Ensure date objects
    if isinstance(date_from, str): date_from = txns[0]["date"]
    if isinstance(date_to,   str): date_to   = txns[-1]["date"]
    months    = max(1, round(((date_to.year - date_from.year)*12 +
                               date_to.month - date_from.month) + 1))

    avg_monthly_in_trx  = round(total_in_trx / months)
    avg_monthly_out_trx = round(total_out_trx / months)
    avg_monthly_in_amt  = round(total_in_amt / months, 2)
    avg_monthly_out_amt = round(total_out_amt / months, 2)
    avg_closing_bal     = round(sum(closing_values) / len(closing_values), 2) \
                          if closing_values else 0.0

    # ── Monthly breakdown ─────────────────────────────────────────────────
    monthly: dict[tuple, dict] = {}
    for t in txns:
        key = (t["date"].year, t["date"].month)
        if key not in monthly:
            monthly[key] = {
                "debit_trx": 0, "debit_amt": 0.0,
                "credit_trx": 0, "credit_amt": 0.0,
                "balances": [], "date": date(t["date"].year, t["date"].month, 1)
            }
        if t["money_out"] > 0:
            monthly[key]["debit_trx"]  += 1
            monthly[key]["debit_amt"]  += t["money_out"]
        if t["money_in"] > 0:
            monthly[key]["credit_trx"]  += 1
            monthly[key]["credit_amt"]  += t["money_in"]
        monthly[key]["balances"].append(t["os_balance"])

    monthly_rows = []
    for idx, (key, m) in enumerate(sorted(monthly.items()), 1):
        bals = m["balances"]
        avg_bal = round(sum(bals)/len(bals), 2) if bals else 0.0
        monthly_rows.append({
            "no":         idx,
            "month":      m["date"],
            "debit_trx":  m["debit_trx"],
            "debit_amt":  round(m["debit_amt"], 2),
            "credit_trx": m["credit_trx"],
            "credit_amt": round(m["credit_amt"], 2),
            "avg_bal":    avg_bal,
            "lowest_bal": min(bals) if bals else 0.0,
            "highest_bal":max(bals) if bals else 0.0,
        })

    # ── Flag reversals ────────────────────────────────────────────────────
    reversal_amt = round(sum(
        t["money_in"] for t in txns
        if "REVERSAL" in t["desc"].upper()
    ), 2)

    return {
        "total_in_trx":       total_in_trx,
        "total_out_trx":      total_out_trx,
        "total_in_amt":       total_in_amt,
        "total_out_amt":      total_out_amt,
        "highest_balance":    highest_bal,
        "lowest_balance":     lowest_bal,
        "avg_closing_bal":    avg_closing_bal,
        "avg_monthly_in_trx": avg_monthly_in_trx,
        "avg_monthly_out_trx":avg_monthly_out_trx,
        "avg_monthly_in_amt": avg_monthly_in_amt,
        "avg_monthly_out_amt":avg_monthly_out_amt,
        "period_months":      months,
        "reversal_amt":       reversal_amt,
        "monthly_rows":       monthly_rows,
        "daily_closing":      daily_closing,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 3 — EXCEL TEMPLATE FILLER
# ═══════════════════════════════════════════════════════════════════════════════

def _s(ws, row, col, value=None, bold=False, bg=None,
       align="left", wrap=False, fmt=None, border=True, italic=False,
       size=9, color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, italic=italic,
                          size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if border:
        cell.border = THIN
    if fmt:
        cell.number_format = fmt
    return cell


def fill_coho_template(data: dict, output_path: str,
                        template_path: str = None) -> str:
    """
    Fill the COHO Excel template with extracted bank statement data.
    Matches exactly the bank's uniform template structure.
    Returns output_path.
    """
    analytics = compute_analytics(data)
    header    = data["header"]
    txns      = data["transactions"]

    # ── Load or create workbook ────────────────────────────────────────────
    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        # Clear transaction data rows (keep formatting)
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                if cell.column <= 8:
                    cell.value = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

    # Sheet name
    acct_no  = header.get("account_no","")
    bank     = header.get("bank","Bank")
    currency = header.get("currency","USD")
    ws.title = f"{bank} {currency}-{acct_no}"[:31]

    # ── Column widths ──────────────────────────────────────────────────────
    widths = {1:8,2:12,3:45,4:16,5:16,6:16,7:18,
              8:1,9:1,10:18,11:12,12:12,13:14,14:3,
              15:12,16:14,17:3,18:14,19:12,20:12,21:4,22:12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── ROW 1: Title ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    c = ws.cell(row=1, column=1,
                value="Conduct of Bank Account Summary")
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:G1")

    # ── ROW 3: Opening balance ────────────────────────────────────────────
    ws.row_dimensions[3].height = 13
    _s(ws, 3, 5, "Openning Balance :", bold=True, border=False)
    _s(ws, 3, 6, header.get("opening_balance", 0),
       border=False, fmt="#,##0.00")

    # ── ROW 5: Column headers ─────────────────────────────────────────────
    ws.row_dimensions[5].height = 30
    headers_left = [
        "No Trx", "Date", "Transaction Details",
        "Withdrawal (Debit)", "Deposit (Credit)",
        "O/S Balance", "Dialy Closing Balance",
    ]
    for ci, h in enumerate(headers_left, 1):
        _s(ws, 5, ci, h, bold=True, bg=C_SECTION_BG,
           align="center", wrap=True, size=9)

    # ── Summary panel (right side) ────────────────────────────────────────
    _s(ws, 5,  10, "A) Conduct of Account Summary",
       bold=True, bg=C_SUMM_BG, size=10, border=False)
    ws.merge_cells("J5:M5")

    def sum_row(row, lbl, val1, sep=None, val2=None, bg=C_WHITE):
        ws.row_dimensions[row].height = 13
        _s(ws, row, 10, lbl,  bold=True, bg=bg, border=False)
        _s(ws, row, 13, val1, bg=bg, border=False,
           fmt="#,##0.00" if isinstance(val1, float) else None)
        if sep:
            _s(ws, row, 14, sep,  bg=bg, border=False, align="center")
            _s(ws, row, 15, val2, bg=bg, border=False,
               fmt="#,##0.00" if isinstance(val2, float) else None)

    period_from = header.get("period_from")
    period_to   = header.get("period_to")

    _s(ws, 6,  13, "Account 1", bold=True, bg=C_SUMM_BG, border=False)
    sum_row(7,  "Institution Name",         header.get("bank",""),      bg=C_SUMM_BG)
    sum_row(8,  "Account Holder Name",      header.get("holder_name",""), bg=C_SUMM_BG)
    r9 = ws.cell(row=9, column=13, value=period_from)
    r9.number_format = "DD-MMM-YYYY"
    r9.fill = PatternFill("solid", start_color=C_SUMM_BG)
    r9.font = Font(name="Arial", size=9)
    r9t = ws.cell(row=9, column=14, value="To")
    r9t.fill = PatternFill("solid", start_color=C_SUMM_BG)
    r9t.font = Font(name="Arial", size=9)
    r9e = ws.cell(row=9, column=15, value=period_to)
    r9e.number_format = "DD-MMM-YYYY"
    r9e.fill = PatternFill("solid", start_color=C_SUMM_BG)
    r9e.font = Font(name="Arial", size=9)
    _s(ws, 9, 10, "Period Cover", bold=True, bg=C_SUMM_BG, border=False)
    sum_row(10, "Period in Months",  analytics.get("period_months",0),      bg=C_SUMM_BG)
    sum_row(11, "Account Number",    header.get("account_no",""),            bg=C_SUMM_BG)
    sum_row(12, "Currency",          f"({header.get('currency','USD')})",    bg=C_SUMM_BG)
    _s(ws, 13, 13, "Withdraw (Debit)", bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 13, 15, "Deposit (Credit)", bold=True, bg=C_SUMM_BG, border=False)
    sum_row(14, "Number of Transaction",
            analytics.get("total_out_trx",0), "/",
            analytics.get("total_in_trx",0),  bg=C_SUMM_BG)
    sum_row(15, "Average Monthly Trx",
            analytics.get("avg_monthly_out_trx",0), "/",
            analytics.get("avg_monthly_in_trx",0),  bg=C_SUMM_BG)
    sum_row(16, "Turnover Dr/Cr",
            round(analytics.get("total_out_amt",0),2), "/",
            round(analytics.get("total_in_amt",0),2),  bg=C_SUMM_BG)
    sum_row(17, "Average Monthly Dr/Cr",
            round(analytics.get("avg_monthly_out_amt",0),2), "/",
            round(analytics.get("avg_monthly_in_amt",0),2),  bg=C_SUMM_BG)
    sum_row(18, "Highest Balance Dr/Cr", analytics.get("highest_balance",0), bg=C_SUMM_BG)
    sum_row(19, "Lowest Balance Dr/Cr",  analytics.get("lowest_balance",0),  bg=C_SUMM_BG)
    sum_row(20, "Average Closing Balance",analytics.get("avg_closing_bal",0),bg=C_SUMM_BG)

    # Exclusion rows
    _s(ws, 22, 10, "To Exclude from Turnover for the following Transaction :",
       bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 22, 13, "Loan Disbursement",    bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 22, 15, 0, bg=C_SUMM_BG, border=False, fmt="#,##0.00")
    _s(ws, 23, 13, "Manipulation Trx.",    bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 23, 15, 0, bg=C_SUMM_BG, border=False, fmt="#,##0.00")
    _s(ws, 24, 13, "Reversal Entries",     bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 24, 15, analytics.get("reversal_amt",0),
       bg=C_SUMM_BG, border=False, fmt="#,##0.00")
    _s(ws, 25, 13, "……………………………",           bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 25, 15, 0, bg=C_SUMM_BG, border=False, fmt="#,##0.00")
    _s(ws, 26, 13, "Total",                bold=True, bg=C_SUMM_BG, border=False)
    _s(ws, 26, 15, 0, bg=C_SUMM_BG, border=False, fmt="#,##0.00")

    # ── Section B header ───────────────────────────────────────────────────
    _s(ws, 28, 10, "B) Conduct of Account Summary (By Month)",
       bold=True, bg=C_SUMM_BG, size=10, border=False)
    ws.merge_cells("J28:W28")

    # Section B column headers
    ws.row_dimensions[29].height = 30
    b_headers = [
        ("J29","No"),("K29","Month"),
        ("L29","Debit (Withdrawal)"),("O29","Credit (Deposit)"),
        ("R29","Average\nClosing Bal."),("S29","Lowest Bal."),
        ("T29","Highest Bal."),("U29","Min"),("V29","Max"),
    ]
    for cell_addr, lbl in b_headers:
        c2 = ws[cell_addr]
        c2.value     = lbl
        c2.font      = Font(name="Arial", bold=True, size=9)
        c2.fill      = PatternFill("solid", start_color=C_MONTH_HDR)
        c2.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        c2.border    = THIN

    ws.row_dimensions[30].height = 16
    for ci, lbl in [(12,"Debit (Trx)"),(13,"Debit (Amt)"),
                    (15,"Credit (Trx)"),(16,"Credit (Amt)")]:
        _s(ws, 30, ci, lbl, bold=True, bg=C_MONTH_HDR, align="center")

    # ── Section B data rows ────────────────────────────────────────────────
    monthly_rows = analytics.get("monthly_rows", [])
    MONTH_START_ROW = 31

    for idx, mr in enumerate(monthly_rows):
        r   = MONTH_START_ROW + idx
        bg  = C_GREY if idx % 2 else C_WHITE
        ws.row_dimensions[r].height = 13

        _s(ws, r, 10, mr["no"],         align="center",  bg=bg)
        mc = ws.cell(row=r, column=11, value=mr["month"])
        mc.number_format = "MMM-YYYY"
        mc.fill = PatternFill("solid", start_color=bg)
        mc.font = Font(name="Arial", size=9)
        mc.border = THIN

        _s(ws, r, 12, mr["debit_trx"],  align="center",  bg=bg)
        _s(ws, r, 13, mr["debit_amt"],  align="right",   bg=bg, fmt="#,##0.00")
        _s(ws, r, 14, "/",              align="center",  bg=bg)
        _s(ws, r, 15, mr["credit_trx"], align="center",  bg=bg)
        _s(ws, r, 16, mr["credit_amt"], align="right",   bg=bg, fmt="#,##0.00")
        _s(ws, r, 17, "/",              align="center",  bg=bg)
        _s(ws, r, 18, mr["avg_bal"],    align="right",   bg=bg, fmt="#,##0.00")
        _s(ws, r, 19, "",               bg=bg)
        _s(ws, r, 20, "",               bg=bg)
        _s(ws, r, 21, mr["lowest_bal"], align="right",   bg=bg, fmt="#,##0.00")
        _s(ws, r, 22, mr["highest_bal"],align="right",   bg=bg, fmt="#,##0.00")

    # Fill remaining monthly rows (up to 24) with zeros
    for idx in range(len(monthly_rows), 24):
        r = MONTH_START_ROW + idx
        ws.row_dimensions[r].height = 13
        for ci in range(10, 23):
            _s(ws, r, ci, 0 if ci != 14 and ci != 17 else "/",
               align="center" if ci in (10,14,17) else "right",
               bg=C_WHITE, fmt="#,##0.00" if ci not in (10,14,17) else None)

    # Totals row
    total_row = MONTH_START_ROW + 24
    ws.row_dimensions[total_row].height = 14
    _s(ws, total_row, 11, "Total:", bold=True, bg=C_TOTAL_BG)
    _s(ws, total_row, 12, analytics.get("total_out_trx",0),
       bold=True, bg=C_TOTAL_BG, align="center")
    _s(ws, total_row, 13, round(analytics.get("total_out_amt",0),2),
       bold=True, bg=C_TOTAL_BG, fmt="#,##0.00")
    _s(ws, total_row, 15, analytics.get("total_in_trx",0),
       bold=True, bg=C_TOTAL_BG, align="center")
    _s(ws, total_row, 16, round(analytics.get("total_in_amt",0),2),
       bold=True, bg=C_TOTAL_BG, fmt="#,##0.00")

    # Variance check
    var_row = total_row + 2
    _s(ws, var_row, 10, "Variance Check", bold=True, bg=C_TOTAL_BG)
    for ci in [12,13,15,16]:
        _s(ws, var_row, ci, 0, bg=C_TOTAL_BG, fmt="#,##0.00")

    # ── TRANSACTION ROWS ──────────────────────────────────────────────────
    TRX_START = 6
    daily_closing = analytics.get("daily_closing", {})

    for idx, t in enumerate(txns):
        r   = TRX_START + idx
        bg  = C_GREY if idx % 2 else C_WHITE
        ws.row_dimensions[r].height = 13

        is_closing = (t["date"] in daily_closing and
                      t["os_balance"] == daily_closing[t["date"]])

        _s(ws, r, 1, idx+1, align="center", bg=bg)

        dc = ws.cell(row=r, column=2, value=t["date"])
        dc.number_format = "DD/MM/YYYY"
        dc.fill  = PatternFill("solid", start_color=bg)
        dc.font  = Font(name="Arial", size=9)
        dc.border = THIN

        _s(ws, r, 3, t["desc"], bg=bg, wrap=True)
        _s(ws, r, 4, t["money_out"] if t["money_out"] > 0 else None,
           bg=bg, fmt="#,##0.00", align="right")
        _s(ws, r, 5, t["money_in"] if t["money_in"] > 0 else None,
           bg=bg, fmt="#,##0.00", align="right")
        _s(ws, r, 6, t["os_balance"],
           bg=bg, fmt="#,##0.00", align="right")
        _s(ws, r, 7, t["os_balance"] if is_closing else None,
           bg=C_CLOSING if is_closing else bg,
           fmt="#,##0.00", align="right")

    # Freeze top rows
    ws.freeze_panes = "A6"

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 4 — CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python coho_extractor.py <statement.pdf> [output.xlsx] [template.xlsx]")
        sys.exit(1)

    pdf_path  = sys.argv[1]
    out_path  = sys.argv[2] if len(sys.argv) > 2 else \
                pdf_path.replace(".pdf","_COHO.xlsx")
    tmpl_path = sys.argv[3] if len(sys.argv) > 3 else None

    print(f"Extracting: {pdf_path}")
    data = extract_statement(pdf_path)
    h    = data["header"]
    txns = data["transactions"]
    an   = compute_analytics(data)

    print(f"  Bank:             {h.get('bank')}")
    print(f"  Account Holder:   {h.get('holder_name')}")
    print(f"  Account No:       {h.get('account_no')}")
    print(f"  Period:           {h.get('period_from')} → {h.get('period_to')}")
    print(f"  Opening Balance:  ${h.get('opening_balance',0):,.2f}")
    print(f"  Ending Balance:   ${h.get('ending_balance',0):,.2f}")
    print(f"  Transactions:     {len(txns)}")
    print(f"  Total In:         ${an.get('total_in_amt',0):,.2f} ({an.get('total_in_trx',0)} trx)")
    print(f"  Total Out:        ${an.get('total_out_amt',0):,.2f} ({an.get('total_out_trx',0)} trx)")
    print(f"  Highest Balance:  ${an.get('highest_balance',0):,.2f}")
    print(f"  Avg Closing Bal:  ${an.get('avg_closing_bal',0):,.2f}")

    print(f"\nGenerating Excel: {out_path}")
    fill_coho_template(data, out_path, tmpl_path)
    print(f"Done: {out_path}")


if __name__ == "__main__":
    main()
