"""
cbc_extractor.py
Extract all data from a Cambodia Credit Bureau (CBC) PDF report
and populate the bank's uniform Excel template.

Usage (standalone):
    python cbc_extractor.py <cbc_report.pdf> [output.xlsx]

Usage (as module):
    from cbc_extractor import extract_cbc_data, fill_excel_template
    data = extract_cbc_data("report.pdf")
    fill_excel_template(data, "output.xlsx")
"""

import re
import io
import os
import sys
import copy
from datetime import datetime
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colours matching the bank template ───────────────────────────────────────
C_GREEN_DARK  = "1F5C2E"   # header bg
C_GREEN_LIGHT = "C6EFCE"   # section bg
C_YELLOW      = "FFFF99"   # totals bg
C_BLUE_LIGHT  = "DCE6F1"   # active header bg
C_GREY        = "F2F2F2"   # alternate row
C_WHITE       = "FFFFFF"
BORDER_THIN   = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 1 — PDF PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class CBCParser:
    """Parse a CBC PDF report into structured Python dicts."""

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.full_text = ""
        self._extract_text()

    def _extract_text(self):
        with pdfplumber.open(self.pdf_path) as pdf:
            parts = []
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=2, y_tolerance=3)
                if text:
                    parts.append(text)
            self.full_text = "\n".join(parts)

    # ── Header ────────────────────────────────────────────────────────────────

    def _parse_header(self) -> dict:
        h = {}
        patterns = {
            "report_date":    r"Report Date\s+([\d/]+)",
            "enquiry_type":   r"Enquiry Type\s+(.+?)(?:\n|Enquiry Number)",
            "enquiry_number": r"Enquiry Number\s+(\d+)",
            "product_type":   r"Product Type\s+(.+?)(?:\n|Number)",
            "num_applicants": r"Number of Applicants\s+(\d+)",
            "account_type":   r"Account Type\s+(\w+)",
            "member_ref":     r"Member Reference\s+(\S+)",
            "amount":         r"Amount\s+((?:USD|KHR)\s+[\d,\.]+)",
        }
        for key, pat in patterns.items():
            m = re.search(pat, self.full_text, re.IGNORECASE)
            h[key] = m.group(1).strip() if m else ""
        return h

    # ── Split applicants ──────────────────────────────────────────────────────

    def _split_applicants(self) -> list[str]:
        """Split full text into per-applicant sections."""
        parts = re.split(r"Applicant \d+ of \d+", self.full_text)
        # If no split occurred, check if it's a single-applicant report
        # (no "Applicant X of Y" marker) — use text after header
        if len(parts) == 1:
            # Single applicant — skip the header section
            # Header ends around "Data Provided vs" or "Applicant Type"
            m = re.search(r"(Data Provided vs|Applicant Type\s+\w+)", self.full_text)
            if m:
                parts = [self.full_text[m.start():]]
            else:
                parts = [self.full_text]
        # For multi-applicant: skip parts[0] (header before first marker)
        # For single-applicant: our fix puts data in parts[0], return it
        if len(parts) == 1:
            return [p.strip() for p in parts if p.strip()]
        return [p.strip() for p in parts[1:] if p.strip()]

    # ── Personal info ─────────────────────────────────────────────────────────

    def _parse_personal(self, text: str) -> dict:
        p = {}
        patterns = {
            "applicant_type":  r"Applicant Type\s+(\w+)",
            "id_type":         r"ID Type and Number\s+(National ID|Passport|Family Book)\s+(\S+)",
            "id_expiry":       r"ID Expiry Date\s+([\d/]+)",
            "name_en_family":  r"Family Name\s+(\w+)\s+\(",
            "name_en_first":   r"First Name\s+(\w+)\s+\(",
            "dob":             r"Date of Birth\s+([\d/]+)",
            "gender":          r"Gender\s+(Male|Female)",
            "marital_status":  r"Marital Status\s+(\w+)",
            "nationality":     r"Nationality\s+(\w+)",
            "province":        r"Place of Birth Province\s+(\w[\w\s]+?)(?:\s*\()",
        }
        for key, pat in patterns.items():
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                p[key] = m.group(1).strip() if m.lastindex == 1 else \
                         (m.group(1).strip() + " " + m.group(2).strip()).strip()
        # ID number
        m = re.search(r"National ID\s+(\d{9,12})", text)
        p["id_number"] = m.group(1) if m else ""
        # Full English name
        fam = p.get("name_en_family", "")
        fst = p.get("name_en_first", "")
        p["full_name_en"] = f"{fam} {fst}".strip()
        return p

    # ── Summary ───────────────────────────────────────────────────────────────

    def _parse_summary(self, text: str) -> dict:
        s = {}
        num_pats = {
            "prev_enquiries":    r"All Previous Enquiries\s+(\d+)",
            "enquiries_30d":     r"Enquiries for Previous 30 Days\s+(\d+)",
            "total_accounts":    r"Total Accounts\s+(\d+)",
            "normal_accounts":   r"Normal Accounts\s+(\d+)",
            "delinquent":        r"Delinquent Accounts\s+(\d+)",
            "closed_accounts":   r"Closed Accounts\s+(\d+)",
            "reject_accounts":   r"Reject Accounts\s+(\d+)",
            "writeoff_accounts": r"Write Off Accounts\s+(\d+)",
            "guaranteed":        r"Guaranteed Accounts\s+(\d+)",
        }
        for key, pat in num_pats.items():
            m = re.search(pat, text)
            s[key] = int(m.group(1)) if m else 0

        # Total limits / liabilities (handle multi-currency)
        limits_usd = re.search(r"Total Limits\s+USD\s+([\d,\.]+)", text)
        limits_khr = re.search(r"Total Limits\s+KHR\s+([\d,\.]+)|KHR\s+([\d,\.]+)\s*Total Liabilities", text)
        liab_usd   = re.search(r"Total Liabilities\s+USD\s+([\d,\.]+)", text)
        liab_khr   = re.search(r"Total Liabilities\s+KHR\s+([\d,\.]+)", text)

        def _num(m, g=1):
            if not m: return 0.0
            try:
                grp = m.group(g) or (m.group(g+1) if m.lastindex >= g+1 else None)
                return float(str(grp).replace(",","")) if grp else 0.0
            except: return 0.0

        s["total_limits_usd"]  = _num(limits_usd)
        s["total_limits_khr"]  = 0.0
        s["total_liab_usd"]    = _num(liab_usd)
        s["total_liab_khr"]    = _num(liab_khr)
        return s

    # ── Account blocks ────────────────────────────────────────────────────────

    def _parse_accounts(self, text: str) -> dict:
        """Extract all account detail blocks from an applicant section."""
        active   = []
        closed   = []
        writeoff = []
        guaranteed_active  = []
        guaranteed_closed  = []

        # Try summary table format first (DD/MM/YYYY LENDER Type REF LoanType CCY Amt Role)
        loan_re = re.compile(
            r"(\d{2}/\d{2}/\d{4})\s+([A-Z][A-Z\s&.\'\-]+?)\s+"
            r"(New|Review|Restructure|Settled|Write Off)\s+"
            r"(Single|Joint|as Single|as Joint)\s+(\S+)\s+"
            r"([\w\s]+?)\s+(USD|KHR)\s*([\d,\.]*)"
            r"\s*(Primary|Guarantor|Co-Borrower|Co-borrower)?",
            re.IGNORECASE
        )
        summary_matches = list(loan_re.finditer(text))
        if summary_matches:
            for m in summary_matches:
                role    = (m.group(9) or "Primary").strip()
                amt_str = m.group(8).replace(",","").strip()
                amt     = float(amt_str) if amt_str else 0.0
                loan = {
                    "date":         m.group(1),
                    "lender":       m.group(2).strip(),
                    "enquiry_type": m.group(3).strip(),
                    "account_type": m.group(4).strip(),
                    "reference":    m.group(5).strip(),
                    "loan_type":    m.group(6).strip(),
                    "currency":     m.group(7).strip(),
                    "loan_amount":  amt,
                    "role":         role,
                    "status":       "Normal",
                }
                if role.lower() in ("guarantor","co-borrower","co borrower"):
                    guaranteed_active.append(loan)
                else:
                    active.append(loan)
            return {
                "accounts":          active,
                "closed_accounts":   closed,
                "writeoff_accounts": writeoff,
                "guaranteed_active": guaranteed_active,
                "guaranteed_closed": guaranteed_closed,
            }

        # Find all "Creditor ..." blocks
        block_starts = [m.start() for m in re.finditer(r"\nCreditor\s+", text)]
        for i, start in enumerate(block_starts):
            end  = block_starts[i+1] if i+1 < len(block_starts) else len(text)
            blob = text[start:end]
            acc  = self._parse_one_account(blob)
            if not acc: continue

            status = acc.get("status","").lower()
            is_guar = acc.get("is_guaranteed", False)

            if is_guar:
                if status == "closed":
                    guaranteed_closed.append(acc)
                else:
                    guaranteed_active.append(acc)
            elif status == "closed":
                closed.append(acc)
            elif status == "normal" or status == "":
                active.append(acc)

        return {
            "active":            active,
            "closed":            closed,
            "writeoff":          writeoff,
            "guaranteed_active": guaranteed_active,
            "guaranteed_closed": guaranteed_closed,
        }

    def _parse_one_account(self, blob: str) -> dict | None:
        """Parse a single account block."""
        a = {}
        pats = {
            "creditor":          r"Creditor\s+(.+?)(?:Issue Date|$)",
            "product_type":      r"Product Type\s+(.+?)(?:Expiry Date|Applicant Type|$)",
            "applicant_type":    r"Applicant Type\s+(\w+)",
            "loan_term_type":    r"Loan Term Type\s+(.+?)(?:\n|Account Number|$)",
            "account_number":    r"Account Number\s+(\S+)",
            "closed_date":       r"Closed Date\s+([\d/]+)",
            "status":            r"Status\s+(\w+)",
            "currency":          r"Currency\s+(US Dollar|Riel|USD|KHR)",
            "payment_frequency": r"Payment Frequency\s+(\w+)",
            "restructured":      r"Restructured Loan\s+(Yes|No)",
            "security_type":     r"Security Type\s+(.+?)(?:\n|As of Date|Outstanding|$)",
            "advisory":          r"Advisory Message\s+(.+?)(?:\n\n|\Z)",
        }
        date_pats = {
            "issue_date":       r"Issue Date\s+([\d/]+)",
            "expiry_date":      r"Expiry Date\s+([\d/]+)",
            "last_payment_date":r"Last Payment Date\s+([\d/]+)",
            "next_payment_date":r"Next Payment Date\s+([\d/]+)",
            "as_of_date":       r"As of Date\s+([\d/]+)",
        }
        num_pats = {
            "limit":            r"Limit\s+([\d,\.]+)",
            "last_amount_paid": r"Last Amount Paid\s+([\d,\.]+)",
            "tenure":           r"Tenure\s+(\d+)",
            "past_due":         r"Past Due\s+([\d,\.]+)",
            "outstanding":      r"Outstanding Balance\s+([\d,\.]+)",
            "installment":      r"Installment Amount\s+([\d,\.]+)",
        }

        for key, pat in pats.items():
            m = re.search(pat, blob, re.IGNORECASE | re.DOTALL)
            a[key] = m.group(1).strip() if m else ""

        for key, pat in date_pats.items():
            m = re.search(pat, blob)
            a[key] = m.group(1).strip() if m else ""

        for key, pat in num_pats.items():
            m = re.search(pat, blob)
            try:
                a[key] = float(m.group(1).replace(",","")) if m else 0.0
            except:
                a[key] = 0.0

        # Normalize currency
        cur = a.get("currency","")
        if "dollar" in cur.lower() or cur.upper() == "USD":
            a["currency"] = "USD"
        elif "riel" in cur.lower() or cur.upper() == "KHR":
            a["currency"] = "KHR"

        # Last 24 cycles
        m = re.search(r"Last 24 Cycles\s+([0-9A-Z QMC]+)", blob)
        a["last_24_cycles"] = m.group(1).strip() if m else ""

        # Guaranteed flag
        a["is_guaranteed"] = (a.get("applicant_type","").lower() == "guarantee")

        if not a.get("creditor"):
            return None
        return a

    # ── Employment ────────────────────────────────────────────────────────────

    def _parse_employment(self, text: str) -> list[dict]:
        jobs = []
        blocks = re.split(r"(?=Employment Status\s+Current)", text)
        for b in blocks:
            if "Employment Status" not in b: continue
            j = {}
            for key, pat in {
                "status":     r"Employment Status\s+(\w+)",
                "employer":   r"Employer\s+(.+?)(?:\n|Occupation)",
                "occupation": r"Occupation\s+(.+?)(?:\n|Employment Type)",
                "emp_type":   r"Employment Type\s+(\w[\w\-]+)",
                "service_m":  r"Length of Service\(M\)\s+(\d+)",
                "income":     r"Total Monthly\s*Salary/Income\s+([\d,\.]+)",
                "currency":   r"Currency\s+(US Dollar|Riel)",
            }.items():
                m = re.search(pat, b, re.IGNORECASE)
                j[key] = m.group(1).strip() if m else ""
            if j.get("status"): jobs.append(j)
        return jobs

    # ── Main parse ────────────────────────────────────────────────────────────

    def parse(self) -> dict:
        header      = self._parse_header()
        applicant_texts = self._split_applicants()
        applicants  = []
        for i, atext in enumerate(applicant_texts):
            personal   = self._parse_personal(atext)
            summary    = self._parse_summary(atext)
            accounts   = self._parse_accounts(atext)
            employment = self._parse_employment(atext)
            applicants.append({
                "index":      i + 1,
                "personal":   personal,
                "summary":    summary,
                "accounts":   accounts,
                "employment": employment,
                "raw_text":   atext,
            })
        return {
            "header":     header,
            "applicants": applicants,
            "parsed_at":  datetime.utcnow().isoformat(),
            "source_file": self.pdf_path,
        }


def extract_cbc_data(pdf_path: str) -> dict:
    """Public function: extract CBC data from PDF."""
    parser = CBCParser(pdf_path)
    return parser.parse()


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 2 — EXCEL TEMPLATE FILLER
# ═══════════════════════════════════════════════════════════════════════════════

EXCHANGE_RATE = 4000   # KHR per USD, default

HEADERS_ACTIVE = [
    "No.", "Account Type", "Status", "Creditor", "Product Type",
    "Account Number", "Loan Term Type", "Issue Date", "Expiry Date",
    "Closed Date", "Currency", "Limit", "Outstanding Balance",
    "Installment Amount", "Last Amount Paid", "Last Payment Date",
    "Next Payment Date", "Tenure\n(m)", "Payment Frequency",
    "Past Due", "Security Type", "Restructured Loan",
    "As of Date", "Last 24 Cycles", "Advisory Message",
]

COL_WIDTHS = {
    1:  6,   # No.
    2:  13,  # Account Type
    3:  10,  # Status
    4:  22,  # Creditor
    5:  20,  # Product Type
    6:  20,  # Account Number
    7:  15,  # Loan Term Type
    8:  11,  # Issue Date
    9:  11,  # Expiry Date
    10: 11,  # Closed Date
    11: 9,   # Currency
    12: 14,  # Limit
    13: 18,  # Outstanding Balance
    14: 16,  # Installment Amount
    15: 16,  # Last Amount Paid
    16: 14,  # Last Payment Date
    17: 14,  # Next Payment Date
    18: 10,  # Tenure
    19: 14,  # Payment Frequency
    20: 10,  # Past Due
    21: 28,  # Security Type
    22: 14,  # Restructured Loan
    23: 11,  # As of Date
    24: 26,  # Last 24 Cycles
    25: 30,  # Advisory Message
}


def _cell_style(ws, row, col, value=None, bold=False, bg=None,
                align="left", wrap=False, fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if border:
        cell.border = BORDER_THIN
    if fmt:
        cell.number_format = fmt
    return cell


def _to_usd(amount: float, currency: str, rate: float = EXCHANGE_RATE) -> float:
    if currency == "KHR":
        return round(amount / rate, 2)
    return amount


def _fill_applicant_sheet(wb: openpyxl.Workbook, applicant: dict,
                           header: dict, sheet_name: str):
    """Fill one sheet of the Excel workbook for one applicant."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    personal = applicant["personal"]
    summary  = applicant["summary"]
    accounts = applicant["accounts"]
    name_en  = personal.get("full_name_en", sheet_name)
    enq_date_raw = header.get("report_date", "")
    try:
        enq_dt   = datetime.strptime(enq_date_raw, "%d/%m/%Y")
        enq_date = enq_dt.strftime("%d-%b-%Y")
    except:
        enq_date = enq_date_raw

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 18
    c = ws.cell(row=1, column=1, value="Credit Bureau Loan Account Listing")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=C_GREEN_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:Y1")

    # ── Row 2: Applicant info label ───────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    c = ws.cell(row=2, column=2, value="Applicant Information")
    c.font = Font(name="Arial", bold=True, size=9)
    c.fill = PatternFill("solid", start_color=C_GREEN_LIGHT)

    # ── Rows 3-9: Applicant info ──────────────────────────────────────────────
    info_rows = [
        ("Applicant Name", name_en,
         "Enquiry No.",    header.get("enquiry_number",""), 3),
        ("Enquiry Date",   enq_date,
         "Exchange Rate USD 1 =", EXCHANGE_RATE, 4),
        ("Primary Accounts",    summary.get("total_accounts",0),
         "Guaranteed Accounts", summary.get("guaranteed",0), 5),
        ("Normal Accounts",     summary.get("normal_accounts",0),
         "Normal Accounts",     0, 6),
        ("Delinquent Accounts", summary.get("delinquent",0),
         "Delinquent Accounts", 0, 7),
        ("Closed Accounts",     summary.get("closed_accounts",0),
         "Closed Accounts",     1, 8),
        ("Write Off Accounts",  summary.get("writeoff_accounts",0),
         "Write Off Accounts",  0, 9),
    ]
    for lbl1, val1, lbl2, val2, r in info_rows:
        ws.row_dimensions[r].height = 13
        _cell_style(ws, r, 2, lbl1,  bold=True, border=False)
        _cell_style(ws, r, 3, val1,  border=False)
        _cell_style(ws, r, 4, lbl2,  bold=True, border=False)
        _cell_style(ws, r, 5, val2,  border=False)

    cur_row = 10

    # ── Helper: write section ─────────────────────────────────────────────────
    def write_section(title: str, accs: list, guar_accs: list,
                      is_closed: bool = False, is_writeoff: bool = False):
        nonlocal cur_row

        ws.row_dimensions[cur_row].height = 14
        c = ws.cell(row=cur_row, column=1, value=title)
        c.font  = Font(name="Arial", bold=True, size=10)
        c.fill  = PatternFill("solid", start_color=C_GREEN_DARK)
        c.font  = Font(name="Arial", bold=True, color="FFFFFF")
        ws.merge_cells(f"A{cur_row}:Y{cur_row}")
        cur_row += 1

        # Header row
        ws.row_dimensions[cur_row].height = 30
        for ci, h in enumerate(HEADERS_ACTIVE, 1):
            _cell_style(ws, cur_row, ci, h, bold=True,
                        bg=C_BLUE_LIGHT, align="center", wrap=True)
        cur_row += 1

        # Write Primary accounts
        c = ws.cell(row=cur_row, column=1, value="Primary Account")
        c.font = Font(name="Arial", bold=True, size=9, italic=True)
        c.fill = PatternFill("solid", start_color=C_GREEN_LIGHT)
        ws.merge_cells(f"A{cur_row}:Y{cur_row}")
        cur_row += 1

        primary_limit_usd     = 0.0
        primary_balance_usd   = 0.0
        primary_install_usd   = 0.0
        primary_last_paid_usd = 0.0

        for idx, acc in enumerate(accs, 1):
            ws.row_dimensions[cur_row].height = 13
            bg = C_GREY if idx % 2 == 0 else C_WHITE
            status = "Closed" if is_closed else acc.get("status","Normal")

            vals = [
                idx,
                "Primary",
                status,
                acc.get("creditor",""),
                acc.get("product_type",""),
                acc.get("account_number",""),
                acc.get("loan_term_type",""),
                acc.get("issue_date",""),
                acc.get("expiry_date",""),
                acc.get("closed_date",""),
                acc.get("currency",""),
                acc.get("limit",0),
                acc.get("outstanding",0),
                acc.get("installment",0),
                acc.get("last_amount_paid",0),
                acc.get("last_payment_date",""),
                acc.get("next_payment_date",""),
                acc.get("tenure",0),
                acc.get("payment_frequency",""),
                acc.get("past_due",0),
                acc.get("security_type",""),
                acc.get("restructured","No"),
                acc.get("as_of_date",""),
                acc.get("last_24_cycles",""),
                acc.get("advisory",""),
            ]
            for ci, v in enumerate(vals, 1):
                fmt = None
                if ci in (12,13,14,15,20):
                    fmt = "#,##0.00"
                _cell_style(ws, cur_row, ci, v, bg=bg, fmt=fmt)

            cur = acc.get("currency","USD")
            primary_limit_usd     += _to_usd(acc.get("limit",0), cur)
            primary_balance_usd   += _to_usd(acc.get("outstanding",0), cur)
            primary_install_usd   += _to_usd(acc.get("installment",0), cur)
            primary_last_paid_usd += _to_usd(acc.get("last_amount_paid",0), cur)
            cur_row += 1

        # Guarantee accounts
        c = ws.cell(row=cur_row, column=1, value="Guarantee Account")
        c.font = Font(name="Arial", bold=True, size=9, italic=True)
        c.fill = PatternFill("solid", start_color=C_GREEN_LIGHT)
        ws.merge_cells(f"A{cur_row}:Y{cur_row}")
        cur_row += 1

        guar_limit_usd     = 0.0
        guar_balance_usd   = 0.0
        guar_install_usd   = 0.0
        guar_last_paid_usd = 0.0

        for idx, acc in enumerate(guar_accs, 1):
            ws.row_dimensions[cur_row].height = 13
            bg = C_GREY if idx % 2 == 0 else C_WHITE
            status = "Closed" if is_closed else acc.get("status","Normal")
            vals = [
                idx, "Guarantee", status,
                acc.get("creditor",""), acc.get("product_type",""),
                acc.get("account_number",""), acc.get("loan_term_type",""),
                acc.get("issue_date",""), acc.get("expiry_date",""),
                acc.get("closed_date",""), acc.get("currency",""),
                acc.get("limit",0), acc.get("outstanding",0),
                acc.get("installment",0), acc.get("last_amount_paid",0),
                acc.get("last_payment_date",""), acc.get("next_payment_date",""),
                acc.get("tenure",0), acc.get("payment_frequency",""),
                acc.get("past_due",0), acc.get("security_type",""),
                acc.get("restructured","No"), acc.get("as_of_date",""),
                acc.get("last_24_cycles",""), acc.get("advisory",""),
            ]
            for ci, v in enumerate(vals, 1):
                fmt = "#,##0.00" if ci in (12,13,14,15,20) else None
                _cell_style(ws, cur_row, ci, v, bg=bg, fmt=fmt)

            cur = acc.get("currency","USD")
            guar_limit_usd     += _to_usd(acc.get("limit",0), cur)
            guar_balance_usd   += _to_usd(acc.get("outstanding",0), cur)
            guar_install_usd   += _to_usd(acc.get("installment",0), cur)
            guar_last_paid_usd += _to_usd(acc.get("last_amount_paid",0), cur)
            cur_row += 1

        if not guar_accs:
            # Empty placeholder rows
            for k in range(1, 3):
                _cell_style(ws, cur_row, 1, k)
                cur_row += 1

        # Totals rows
        ws.row_dimensions[cur_row].height = 13
        totals_primary = [
            "Total Primary (in USD) :", None, None, None, None, None,
            None, None, None, None, None,
            round(primary_limit_usd,2),
            round(primary_balance_usd,2),
            round(primary_install_usd,2),
            round(primary_last_paid_usd,2),
        ]
        for ci, v in enumerate(totals_primary, 1):
            c = _cell_style(ws, cur_row, ci, v, bold=True, bg=C_YELLOW,
                            fmt="#,##0.00" if ci >= 12 else None)
        ws.merge_cells(f"A{cur_row}:K{cur_row}")
        cur_row += 1

        ws.row_dimensions[cur_row].height = 13
        totals_guar = [
            "Total Guarantee (in USD) :", None, None, None, None, None,
            None, None, None, None, None,
            round(guar_limit_usd,2),
            round(guar_balance_usd,2),
            round(guar_install_usd,2),
            round(guar_last_paid_usd,2),
        ]
        for ci, v in enumerate(totals_guar, 1):
            _cell_style(ws, cur_row, ci, v, bold=True, bg=C_YELLOW,
                        fmt="#,##0.00" if ci >= 12 else None)
        ws.merge_cells(f"A{cur_row}:K{cur_row}")
        cur_row += 2   # blank line

    # ── Write all three sections ──────────────────────────────────────────────
    write_section(
        "I. Active Accounts",
        accounts.get("active", []),
        accounts.get("guaranteed_active", []),
        is_closed=False,
    )
    write_section(
        "II. Closed Accounts",
        accounts.get("closed", []),
        accounts.get("guaranteed_closed", []),
        is_closed=True,
    )
    write_section(
        "III. Write-Off Accounts",
        accounts.get("writeoff", []),
        [],
        is_writeoff=True,
    )

    # Freeze top rows
    ws.freeze_panes = "A13"


def fill_excel_template(data: dict, output_path: str,
                         template_path: str = None) -> str:
    """
    Create a filled Excel workbook from parsed CBC data.
    Returns the output path.
    """
    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        # Clear data but keep at least one sheet visible
        for sname in list(wb.sheetnames)[1:]:
            del wb[sname]
        if wb.worksheets:
            wb.worksheets[0].sheet_state = 'visible'
            wb.active = wb.worksheets[0]
    else:
        wb = openpyxl.Workbook()
        # Keep default sheet as placeholder

    header     = data.get("header", {})
    applicants = data.get("applicants", [])

    for app in applicants:
        personal  = app.get("personal", {})
        name_en   = personal.get("full_name_en","")
        if not name_en:
            name_en = f"Applicant {app['index']}"
        # Sheet name max 31 chars
        sheet_name = name_en[:28] + " (COHO)" if len(name_en) <= 21 \
                     else name_en[:24] + " (COHO)"
        sheet_name = sheet_name[:31]
        _fill_applicant_sheet(wb, app, header, sheet_name)

    # Ensure at least one sheet is visible
    visible_found = False
    for ws in wb.worksheets:
        if ws.sheet_state != 'hidden':
            visible_found = True
            break
    if not visible_found:
        wb.worksheets[0].sheet_state = 'visible'
        wb.active = wb.worksheets[0]
    # Ensure at least one sheet is visible before saving
    has_visible = any(ws.sheet_state != 'hidden' for ws in wb.worksheets)
    if not has_visible and wb.worksheets:
        wb.worksheets[0].sheet_state = 'visible'
        wb.active = wb.worksheets[0]
    # Ensure at least one sheet is visible before saving
    if not wb.worksheets:
        wb.create_sheet("Summary")
    has_visible = any(ws.sheet_state != 'hidden' for ws in wb.worksheets)
    if not has_visible:
        wb.worksheets[0].sheet_state = 'visible'
        wb.active = wb.worksheets[0]
    elif wb.active is None or wb.active.sheet_state == 'hidden':
        for ws in wb.worksheets:
            if ws.sheet_state != 'hidden':
                wb.active = ws
                break
    # Ensure at least one sheet visible
    if not wb.worksheets:
        wb.create_sheet("Summary")
    if not any(ws.sheet_state != 'hidden' for ws in wb.worksheets):
        wb.worksheets[0].sheet_state = 'visible'
        wb.active = wb.worksheets[0]
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 3 — CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python cbc_extractor.py <cbc_report.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else \
               pdf_path.replace(".pdf","_CBC_Summary.xlsx")

    print(f"Extracting: {pdf_path}")
    data = extract_cbc_data(pdf_path)

    h = data["header"]
    print(f"  Report date    : {h.get('report_date')}")
    print(f"  Enquiry No.    : {h.get('enquiry_number')}")
    print(f"  Applicants     : {len(data['applicants'])}")
    for app in data["applicants"]:
        p  = app["personal"]
        s  = app["summary"]
        ac = app["accounts"]
        print(f"\n  Applicant {app['index']}: {p.get('full_name_en','?')}")
        print(f"    Total accounts : {s.get('total_accounts',0)}")
        print(f"    Normal/Closed  : {s.get('normal_accounts',0)} / {s.get('closed_accounts',0)}")
        print(f"    Active found   : {len(ac.get('active',[]))}")
        print(f"    Closed found   : {len(ac.get('closed',[]))}")
        print(f"    Guaranteed     : {len(ac.get('guaranteed_closed',[]))}")

    print(f"\nGenerating Excel: {out_path}")
    fill_excel_template(data, out_path)
    print(f"Done: {out_path}")


if __name__ == "__main__":
    main()
